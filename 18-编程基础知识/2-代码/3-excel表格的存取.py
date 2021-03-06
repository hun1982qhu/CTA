import xlrd
import xlwt
import scipy.stats as stats
import talib as ta
import math
import pandas as pd
import numpy as np
from functools import reduce
from openpyxl.workbook import Workbook
from timeit import default_timer as timer
import datetime
from datetime import datetime
from dateutil.parser import parse
import csv
import tushare as ts
import matplotlib.pyplot as plt
import matplotlib as mpl
from matplotlib import style

style.use('ggplot')
import seaborn as sns

sns.set()

mpl.rcParams['font.sans-serif'] = ['Arial Unicode MS']
mpl.rcParams['axes.unicode_minus'] = False

wb = Workbook()
sheet = wb.active
sheet.title = '关于如何读取excel'
sheet['A1'] = '漫威宇宙'
rows = [['宫崎骏', '久石让'], ['龙猫', '千与千寻']]
for i in rows:
    sheet.append(i)
wb.save('/Users/huangning/PycharmProjects/1-Python学习/1-Python练习/4-爬虫课程/动画片.xlsx')
# excel表格的存取
# （一）使用openpyx模块
# 1. 安装模块
# windows电脑：在终端输入命令 pip install openpyxl
# mac电脑：在终端输入命令pip3 install openpyxl
# 2. 安装好openpyxl模块后，首先要引用它，然后通过openpyxl.Workbook()函数就可以创建新的工作簿，代码如下：
# 引用openpyxl     
import openpyxl 
# 利用openpyxl.Workbook()函数创建新的workbook（工作簿）对象，就是创建新的空的Excel文件。
wb = openpyxl.Workbook()
# 3. 创建完新的工作簿后，还得获取工作表。不然程序会无所适从，不知道要把内容写入哪张工作表里。
# wb.active就是获取这个工作簿的活动表，通常就是第一个工作表。
sheet = wb.active
# 可以用.title给工作表重命名。现在第一个工作表的名称就会由原来默认的“sheet1”改为"new title"。
sheet.title = 'new title’
# 4. 添加完工作表，我们就能来操作单元格，往单元格里写入内容。
# 把'漫威宇宙'赋值给第一个工作表的A1单元格，就是往A1的单元格中写入了'漫威宇宙'。
sheet['A1'] = '漫威宇宙' 
# 如果我们想往工作表里写入一行内容的话，就得用到append函数。
# 把我们想写入的一行内容写成列表，赋值给row。
row = ['美国队长','钢铁侠','蜘蛛侠']
# 用sheet.append()就能往表格里添加这一行文字。
sheet.append(row)
# 如果我们想要一次性写入的不止一行，而是多行内容，又该怎么办？# 先把要写入的多行内容写成列表，再放进大列表里，赋值给rows。
rows = [['美国队长','钢铁侠','蜘蛛侠'],['是','漫威','宇宙', '经典','人物']]
# 遍历rows，同时把遍历的内容添加到表格里，这样就实现了多行写入。
for i in rows:
    sheet.append(i)

# 打印rows
print(rows)
# 成功写入后，我们千万要记得保存这个Excel文件，不然就白写啦！
# 保存新建的Excel文件，并命名为“Marvel.xlsx”
wb.save(‘Marvel.xlsx')
# 5. 读取excel表格
# 读取的代码：
wb = openpyxl.load_workbook('Marvel.xlsx')
sheet = wb['new title']
sheetname = wb.sheetnames
print(sheetname)
A1_cell = sheet['A1']
A1_value = A1_cell.value
print(A1_value)

# openpyxl模块的官方文档：
# https://openpyxl.readthedocs.io/en/stable/

# （二）用pandas命令
# 通常使用xlrd库来读取Excel表格中的数据，用xlwt库将处理后的数据保存为Excel文件。但更一般的情况下，使用pandas库的read_xxx函数（常用的诸如read_excel、read_csv、read_sql_query等）来读取数据，使用to_excel(filename,sheetname)  来导出数据到excel中。
# 主要记住使用pandas命令来存取excel即可
# 1.读取
import pandas as pd
table = pd.read_table('http:\\somelink.csv')  # 读取任一结构型文本数据
csv = pd.read_csv(r'c:\data.csv')  # 读取csv文件
excel = pd.read_excel(r'c:\data.xlsx')  # 读取Excel文件
# 2.存储
writer = pd.ExcelWriter(filename)  # 定义一个向Excel写入数据的对象
df1.to_excel(writer,'Data1')  # 向该Excel中写入df1到Data1这个sheet
df2.to_excel(writer,'Data2')  # 向该Excel中写入df2到Data2这个sheet
writer.save()  # 保存Excel表格
writer.close()  # 关闭Excel表格
# 详细的可以参见网址
# https://zhuanlan.zhihu.com/p/33085521

1. 安装模块
windows电脑：在终端输入命令 pip install openpyxl
mac电脑：在终端输入命令pip3 install openpyxl
2. 安装好openpyxl模块后，首先要引用它，然后通过openpyxl.Workbook()函数就可以创建新的工作簿，代码如下：
# 引用openpyxl     
import openpyxl 
# 利用openpyxl.Workbook()函数创建新的workbook（工作簿）对象，就是创建新的空的Excel文件。
wb = openpyxl.Workbook()
3. 创建完新的工作簿后，还得获取工作表。不然程序会无所适从，不知道要把内容写入哪张工作表里。
# wb.active就是获取这个工作簿的活动表，通常就是第一个工作表。
sheet = wb.active
# 可以用.title给工作表重命名。现在第一个工作表的名称就会由原来默认的“sheet1”改为"new title"。
sheet.title = 'new title’
4. 添加完工作表，我们就能来操作单元格，往单元格里写入内容。
# 把'漫威宇宙'赋值给第一个工作表的A1单元格，就是往A1的单元格中写入了'漫威宇宙'。
sheet['A1'] = '漫威宇宙' 
如果我们想往工作表里写入一行内容的话，就得用到append函数。
# 把我们想写入的一行内容写成列表，赋值给row。
row = ['美国队长','钢铁侠','蜘蛛侠']
# 用sheet.append()就能往表格里添加这一行文字。
sheet.append(row)
如果我们想要一次性写入的不止一行，而是多行内容，又该怎么办？# 先把要写入的多行内容写成列表，再放进大列表里，赋值给rows。
rows = [['美国队长','钢铁侠','蜘蛛侠'],['是','漫威','宇宙', '经典','人物']]
# 遍历rows，同时把遍历的内容添加到表格里，这样就实现了多行写入。
for i in rows:
    sheet.append(i)

# 打印rows
print(rows)
成功写入后，我们千万要记得保存这个Excel文件，不然就白写啦！
# 保存新建的Excel文件，并命名为“Marvel.xlsx”
wb.save(‘Marvel.xlsx')
5. 读取excel表格
# 读取的代码：
wb = openpyxl.load_workbook('Marvel.xlsx')
sheet = wb['new title']
sheetname = wb.sheetnames
print(sheetname)
A1_cell = sheet['A1']
A1_value = A1_cell.value
print(A1_value)

openpyxl模块的官方文档：
https://openpyxl.readthedocs.io/en/stable/

（二）用pandas命令
通常使用xlrd库来读取Excel表格中的数据，用xlwt库将处理后的数据保存为Excel文件。但更一般的情况下，使用pandas库的read_xxx函数（常用的诸如read_excel、read_csv、read_sql_query等）来读取数据，使用to_excel(filename,sheetname)  来导出数据到excel中。
主要记住使用pandas命令来存取excel即可
1.读取
import pandas as pd
table = pd.read_table('http:\\somelink.csv')  # 读取任一结构型文本数据
csv = pd.read_csv(r'c:\data.csv')  # 读取csv文件
excel = pd.read_excel(r'c:\data.xlsx')  # 读取Excel文件
2.存储
writer = pd.ExcelWriter(filename)  # 定义一个向Excel写入数据的对象
df1.to_excel(writer,'Data1')  # 向该Excel中写入df1到Data1这个sheet
df2.to_excel(writer,'Data2')  # 向该Excel中写入df2到Data2这个sheet
writer.save()  # 保存Excel表格
writer.close()  # 关闭Excel表格
详细的可以参见网址
https://zhuanlan.zhihu.com/p/33085521
