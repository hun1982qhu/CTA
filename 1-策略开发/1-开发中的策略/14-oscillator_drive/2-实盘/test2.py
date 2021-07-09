#%%
import openpyxl
import pprint
from openpyxl.utils import get_column_letter, column_index_from_string
import datetime

print("opening workbook...")
wb = openpyxl.load_workbook("C:/Users/huangning/Documents/CTA/1-策略开发/1-开发中的策略/14-oscillator_drive/2-实盘/censuspopdata.xlsx")

sheet = wb["Population by Census Tract"]

countyData = {}

print("reading rows...")

countyData.setdefault(state, {})
countyData[state].setdefault(county, {"tracts": 0, "pop": 0})

for row in range(2, sheet.max_row +1):
    state = sheet["B"+str(row)].value
    county = sheet["C"+str(row)].value
    pop = sheet["D"+str(row)].value  
    countyData[state][county]["tracts"] += 1
    countyData[state][county]["pop"] += int(pop)

